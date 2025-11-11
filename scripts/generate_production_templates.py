#!/usr/bin/env python3
"""
Generate three bilingual (AR/EN) Excel templates for film/production projects:
- Expenses/Budget sheet
- Shot List
- Call Sheet

Outputs files under templates/:
  - ExpensesTemplate.xlsx
  - ShotListTemplate.xlsx
  - CallSheetTemplate.xlsx

Re-run any time to regenerate.
"""
from __future__ import annotations

import os
from datetime import date
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


TEMPLATES_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "templates")


def ensure_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)


# ---------- Generic styling helpers ----------
HEADER_FILL = PatternFill("solid", fgColor="EDF2F7")
# Light blue section shade used across first-page headings (match UI)
LIGHT_BLUE_FILL = PatternFill("solid", fgColor="E0ECFF")
# Solid blue title bar for key section headers (match Crew/Cast headings in UI)
SOLID_BLUE_FILL = PatternFill("solid", fgColor="2563EB")
TITLE_FILL = PatternFill("solid", fgColor="1F2937")  # slate-800
TITLE_FONT = Font(bold=True, color="FFFFFF", size=14)
HEADER_FONT = Font(bold=True, color="000000")
SECTION_FILL_BLUE = PatternFill("solid", fgColor="E0ECFF")
SECTION_FILL_GREEN = PatternFill("solid", fgColor="E6FFFA")
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def style_title(ws, cell_range: str, text: str):
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(":")[0]]
    c.value = text
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")


def style_headers(ws, row: int, headers: List[str]):
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_headers_blue(ws, row: int, headers: List[str]):
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=idx, value=h)
        # Dark text for table headers; keep default header font weight
        cell.font = HEADER_FONT
        cell.fill = SOLID_BLUE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def set_col_widths(ws, widths: List[float]):
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


# ---------- Expenses Template ----------
def build_expenses_template(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"

    # Title and meta
    style_title(ws, "A1:S1", "Expenses Sheet / جدول المصاريف")
    meta_labels = [
        "Production Co.", "Project Title", "Client", "Budget Date", "Prepared by", "Locations",
    ]
    meta_values = ["", "", "", date.today().isoformat(), "", ""]
    for i, (lab, val) in enumerate(zip(meta_labels, meta_values), start=1):
        ws.cell(row=2, column=i, value=f"{lab}")
        ws.cell(row=3, column=i, value=val)
        ws.cell(row=2, column=i).font = Font(bold=True)

    headers = [
        "Code / الكود",
        "Section / القسم",
        "Item / البند",
        "Qty / الكمية",
        "Unit / الوحدة",
        "Rate / السعر",
        "Currency / العملة",
        "Tax % / الضريبة%",
        "Discount / الخصم",
        "Line Total / إجمالي البند",
        "Vendor / المورد",
        "Booking Ref / مرجع الحجز",
        "PO # / أمر الشراء",
        "Status / الحالة",
        "Paid / المدفوع",
        "Balance / المتبقي",
        "Due Date / تاريخ الاستحقاق",
        "Payment Method / طريقة الدفع",
        "Notes / ملاحظات",
    ]
    header_row = 5
    style_headers(ws, header_row, headers)

    set_col_widths(ws, [10, 20, 28, 8, 10, 12, 10, 10, 10, 16, 16, 16, 12, 14, 14, 14, 14, 16, 28])

    # Freeze header
    ws.freeze_panes = f"A{header_row+1}"

    # Dropdowns
    currency_dv = DataValidation(type="list", formula1='"SAR,USD,AED,EUR"', allow_blank=True)
    status_dv = DataValidation(type="list", formula1='"Planned,Pending,Approved,Paid,Partially Paid,Cancelled"', allow_blank=True)
    paymethod_dv = DataValidation(type="list", formula1='"Bank,Card,Cash,Transfer,Other"', allow_blank=True)
    ws.add_data_validation(currency_dv)
    ws.add_data_validation(status_dv)
    ws.add_data_validation(paymethod_dv)

    # Data rows with formulas
    start = header_row + 1
    end = start + 300  # 300 rows to start with
    for r in range(start, end + 1):
        # Borders for the range
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER

        # Line total: (Qty * Rate) * (1 + Tax%) - Discount
        qty = f"D{r}"
        rate = f"F{r}"
        tax = f"H{r}"
        disc = f"I{r}"
        ws[f"J{r}"] = f"=ROUND(({qty}*{rate})*(1+IFERROR({tax},0)/100)-IFERROR({disc},0),2)"
        # Balance: Line Total - Paid
        line_total = f"J{r}"
        paid = f"O{r}"
        ws[f"P{r}"] = f"=IFERROR({line_total}-IFERROR({paid},0), {line_total})"

        # Apply dropdowns
        currency_dv.add(ws[f"G{r}"])
        status_dv.add(ws[f"N{r}"])
        paymethod_dv.add(ws[f"R{r}"])

    # Add example section labels to guide grouping
    example_sections = [
        (start, "Above the Line"),
        (start + 20, "Production Expenses"),
        (start + 120, "Post-Production"),
    ]
    for rr, sec in example_sections:
        ws[f"B{rr}"] = sec
        ws[f"B{rr}"].fill = SECTION_FILL_BLUE
        ws[f"B{rr}"].font = Font(bold=True)

    # Summary sheet
    sum_ws = wb.create_sheet("Summary")
    style_title(sum_ws, "A1:D1", "Budget Summary / ملخص الميزانية")
    sum_headers = ["Section / القسم", "Planned Total", "Paid", "Balance"]
    style_headers(sum_ws, 3, sum_headers)
    set_col_widths(sum_ws, [28, 18, 18, 18])

    sections = ["Above the Line", "Production Expenses", "Post-Production", "Contingency", "Misc"]
    for idx, sec in enumerate(sections, start=4):
        sum_ws[f"A{idx}"] = sec
        # Totals via SUMIF over Expenses!B (Section) and J (Line Total) / O (Paid)
        sum_ws[f"B{idx}"] = f"=SUMIF(Expenses!B:B, \"{sec}\", Expenses!J:J)"
        sum_ws[f"C{idx}"] = f"=SUMIF(Expenses!B:B, \"{sec}\", Expenses!O:O)"
        sum_ws[f"D{idx}"] = f"=B{idx}-C{idx}"

    # Grand totals
    last = 4 + len(sections)
    sum_ws[f"A{last}"] = "Grand Total / الإجمالي"
    sum_ws[f"A{last}"].font = Font(bold=True)
    sum_ws[f"B{last}"] = f"=SUM(B4:B{last-1})"
    sum_ws[f"C{last}"] = f"=SUM(C4:C{last-1})"
    sum_ws[f"D{last}"] = f"=SUM(D4:D{last-1})"

    # Save
    wb.save(path)


# ---------- Shot List Template ----------
def build_shotlist_template(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Shot List"

    style_title(ws, "A1:V1", "Shot List / قائمة اللقطات")

    headers = [
        "#",
        "Scene / المشهد",
        "Description / الوصف",
        "Lens / عدسة",
        "Camera Move / حركة كاميرا",
        "Rig / المثبت",
        "FPS",
        "Res / الدقة",
        "INT/EXT",
        "Day/Night",
        "Location / الموقع",
        "Audio / الصوت",
        "Talent / الممثلين",
        "Props / معدات",
        "Wardrobe / الملابس",
        "VFX Notes / ملاحظات VFX",
        "Safety / السلامة",
        "Est. Time / المدة التقديرية",
        "Priority / الأولوية",
        "Status / الحالة",
        "Notes / ملاحظات",
        "Ref Link / رابط مرجع",
    ]
    header_row = 3
    style_headers(ws, header_row, headers)
    set_col_widths(ws, [4, 16, 28, 10, 16, 12, 6, 8, 10, 10, 16, 14, 18, 12, 12, 16, 12, 14, 12, 12, 24])
    ws.freeze_panes = f"A{header_row+1}"

    # Data validations
    int_ext = DataValidation(type="list", formula1='"INT,EXT"', allow_blank=True)
    day_night = DataValidation(type="list", formula1='"Day,Night"', allow_blank=True)
    priority = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
    status = DataValidation(type="list", formula1='"Planned,Shot,Dropped,Hold"', allow_blank=True)
    ws.add_data_validation(int_ext)
    ws.add_data_validation(day_night)
    ws.add_data_validation(priority)
    ws.add_data_validation(status)

    start = header_row + 1
    end = start + 400
    for r in range(start, end + 1):
        int_ext.add(ws[f"I{r}"])
        day_night.add(ws[f"J{r}"])
        priority.add(ws[f"S{r}"])
        status.add(ws[f"T{r}"])
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER

    wb.save(path)


# ---------- Call Sheet Template ----------
def build_callsheet_template(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Call Sheet"

    # Overall title
    style_title(ws, "A1:L1", "CALL SHEET / نموذج الكول شيت")

    set_col_widths(ws, [18, 18, 18, 16, 16, 18, 14, 14, 14, 14, 18, 22])

    bold = Font(bold=True)

    # Header/meta blocks
    headers = [
        (2, ["Production", "Project Title", "Client", "Date", "Shoot Days", "Location City"]),
        (4, ["Producer", "Director", "DOP", "1st AD", "PM", "Contact (Tel/Email)"]),
        (6, ["Call Time", "Ready to Shoot", "Lunch", "Est. Wrap", "Sunrise", "Sunset"]),
        (8, ["Weather", "Temp", "Wind", "Rain %", "Nearest Hospital", "Map Link"]),
    ]
    for start_row, labels in headers:
        for i, lab in enumerate(labels, start=1):
            # Label row with light-blue shading
            lab_cell = ws.cell(row=start_row, column=i, value=lab)
            lab_cell.font = bold
            lab_cell.fill = LIGHT_BLUE_FILL
            # Value row under it
            ws.cell(row=start_row + 1, column=i, value="")

    # Set "Date" (row 3, column 4) to today's date by default
    try:
        ws.cell(row=3, column=4, value=date.today().isoformat())
    except Exception:
        pass

    # Important notes (full width)
    ws.merge_cells("A10:L10")
    ws["A10"].value = "Important Notes / ملاحظات مهمة"
    ws["A10"].font = Font(bold=True)
    ws["A10"].fill = LIGHT_BLUE_FILL
    ws.merge_cells("A11:L14")
    ws["A11"].alignment = Alignment(wrap_text=True, vertical="top")

    # Cast calls table
    ws["A16"].value = "Cast Calls / جدول مواعيد الممثلين"
    ws["A16"].font = Font(bold=True, color="FFFFFF")
    ws["A16"].fill = SOLID_BLUE_FILL
    cast_headers = ["Name", "Role", "Call", "Makeup", "Wardrobe", "On Set", "Notes"]
    # Cast table headers (row 17): keep standard header background
    style_headers(ws, 17, cast_headers)
    for r in range(18, 36):  # 18 rows
        for c in range(1, len(cast_headers) + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER

    # Key contacts block
    base_r = 16
    col = 9
    kc = ws.cell(row=base_r, column=col, value="Key Contacts / الأسماء المهمة")
    kc.font = bold
    kc.fill = LIGHT_BLUE_FILL
    contacts = ["Producer", "Director", "1st AD", "DOP", "Gaffer", "Sound", "Wardrobe", "Makeup", "PA", "Driver"]
    r = base_r + 1
    for role in contacts:
        ws.cell(row=r, column=col, value=role).font = Font(bold=True)
        ws.cell(row=r, column=col + 1, value="Name / Phone / Radio")
        r += 1

    # Schedule table
    sched_start = 38
    ws[f"A{sched_start-1}"].value = "Schedule / جدول اليوم"
    ws[f"A{sched_start-1}"].font = bold
    sched_headers = [
        "Time (Duration)",
        "Shot #",
        "Description",
        "Location",
        "MOVEMENT",
        "VO",
        "Cast",
        "Action Props",
        "Notes",
    ]
    style_headers(ws, sched_start, sched_headers)
    for r in range(sched_start + 1, sched_start + 1 + 40):
        for c in range(1, len(sched_headers) + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER

    # Wrap time
    wrap_row = sched_start + 42
    ws.merge_cells(f"A{wrap_row}:I{wrap_row}")
    ws[f"A{wrap_row}"] = "WRAP / نهاية اليوم"
    ws[f"A{wrap_row}"].alignment = Alignment(horizontal="center")
    ws[f"A{wrap_row}"].font = Font(bold=True)

    # Freeze panes just under meta
    ws.freeze_panes = "A18"

    wb.save(path)


def main():
    ensure_dir(TEMPLATES_DIR)
    build_expenses_template(os.path.join(TEMPLATES_DIR, "ExpensesTemplate.xlsx"))
    build_shotlist_template(os.path.join(TEMPLATES_DIR, "ShotListTemplate.xlsx"))
    build_callsheet_template(os.path.join(TEMPLATES_DIR, "CallSheetTemplate.xlsx"))
    print(f"Templates generated in: {TEMPLATES_DIR}")


if __name__ == "__main__":
    main()
